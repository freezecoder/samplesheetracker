"""
xlsm extractor
--------------
Reads an Excel macro-enabled workbook (.xlsm/.xlsx) and returns two DataFrames:
  [0] – the table matching a label (default "E.1. Table 5") in the E1 Norm+Pool sheet
  [1] – the full H.1 Manifest sheet

Usage
-----
As a module:
    from extractor import extract
    e1_df, manifest_df = extract("path/to/file.xlsm")

    # Override sheet names or table label:
    e1_df, manifest_df = extract(
        "path/to/file.xlsm",
        e1_sheet="My Sheet",
        manifest_sheet="My Manifest",
        table_label="E.1. Table 5",
        no_header=False,
    )

As a standalone script:
    python extractor.py file.xlsm
    python extractor.py file.xlsm --out-dir ./output
    python extractor.py file.xlsm --format csv
    python extractor.py file.xlsm --e1-sheet "E1 Norm+Pool" --table-label "E.1. Table 5"
    python extractor.py file.xlsm --list-sheets
    python extractor.py file.xlsm --show e1
    python extractor.py file.xlsm --show manifest --max-rows 20
"""

from __future__ import annotations

import sys
import argparse
import pathlib
from typing import List, Optional
import openpyxl
import pandas as pd


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def _find_table_bounds(ws, label: str):
    """
    Scan *ws* for a cell containing *label* (case-insensitive substring match).
    Returns (header_row, col_start, col_end, data_start_row, data_end_row).

    Layout assumed:
        [label cell]          ← label_row
        [col1] [col2] ...     ← header_row  (first non-empty row after label)
        [data] [data]  ...    ← data rows
        <empty row>           ← signals end of table
    """
    label_lower = label.strip().lower()
    label_row = None

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and label_lower in cell.value.strip().lower():
                label_row = cell.row
                break
        if label_row:
            break

    if label_row is None:
        raise ValueError(
            f"Could not find a cell matching '{label}' in sheet '{ws.title}'."
        )

    # First non-empty row after the label → header
    header_row = label_row + 1
    while header_row <= ws.max_row:
        if any(ws.cell(row=header_row, column=c).value is not None
               for c in range(1, ws.max_column + 1)):
            break
        header_row += 1

    # Column extent from the header row
    col_start = col_end = None
    for c in range(1, ws.max_column + 1):
        if ws.cell(row=header_row, column=c).value is not None:
            if col_start is None:
                col_start = c
            col_end = c

    if col_start is None:
        raise ValueError(
            f"Header row {header_row} is empty – cannot determine table columns."
        )

    # Data rows until first fully-empty row
    data_start = header_row + 1
    data_end = data_start - 1
    for r in range(data_start, ws.max_row + 1):
        if all(ws.cell(row=r, column=c).value is None
               for c in range(col_start, col_end + 1)):
            break
        data_end = r

    return header_row, col_start, col_end, data_start, data_end


def _ws_to_dataframe(ws, min_row=1, min_col=1, max_row=None, max_col=None,
                     has_header=True):
    """Read a rectangular region of *ws* into a DataFrame."""
    max_row = max_row or ws.max_row
    max_col = max_col or ws.max_column

    data = [
        [ws.cell(row=r, column=c).value for c in range(min_col, max_col + 1)]
        for r in range(min_row, max_row + 1)
    ]

    if not data:
        return pd.DataFrame()

    if has_header:
        seen: dict[str, int] = {}
        headers = []
        for h in data[0]:
            h = str(h) if h is not None else "Unnamed"
            n = seen.get(h, 0)
            headers.append(h if n == 0 else f"{h}.{n}")
            seen[h] = n + 1
        return pd.DataFrame(data[1:], columns=headers)

    return pd.DataFrame(data)


def _resolve_sheet(wb, candidates: List[str], param_name: str) -> str:
    """Return the first sheetname that case-insensitively contains any candidate."""
    for name in wb.sheetnames:
        if any(c.lower() in name.lower() for c in candidates):
            return name
    raise ValueError(
        f"No sheet matching {candidates!r} found ({param_name}). "
        f"Available sheets: {wb.sheetnames}"
    )


# ---------------------------------------------------------------------------
# public API
# ---------------------------------------------------------------------------

def extract(
    filepath: str,
    *,
    e1_sheet: Optional[str] = None,
    manifest_sheet: Optional[str] = None,
    table_label: str = "E.1. Table 5",
    no_header: bool = False,
) -> List[pd.DataFrame]:
    """
    Parse *filepath* and return ``[e1_table_df, manifest_df]``.

    Parameters
    ----------
    filepath : str
        Path to the .xlsm / .xlsx workbook.
    e1_sheet : str, optional
        Exact sheet name for the E1 Norm+Pool sheet.  If omitted the sheet is
        located by a case-insensitive search for "e1 norm".
    manifest_sheet : str, optional
        Exact sheet name for the manifest sheet.  If omitted the sheet is
        located by a case-insensitive search for "manifest".
    table_label : str
        Substring to search for when locating the target table.
        Default: ``"E.1. Table 5"``.
    no_header : bool
        When *True*, treat the first data row as data (no column names).

    Returns
    -------
    list[pd.DataFrame]
        ``[e1_df, manifest_df]``
    """
    wb = openpyxl.load_workbook(filepath, data_only=True, keep_vba=False)

    # ---- Sheet 1: E1 Norm+Pool ----------------------------------------
    e1_name = e1_sheet or _resolve_sheet(wb, ["e1 norm", "e1norm"], "--e1-sheet")
    ws_e1 = wb[e1_name]
    hdr_row, col_s, col_e, _, data_e = _find_table_bounds(ws_e1, table_label)
    e1_df = _ws_to_dataframe(
        ws_e1,
        min_row=hdr_row,
        min_col=col_s,
        max_row=data_e,
        max_col=col_e,
        has_header=not no_header,
    )

    # ---- Sheet 2: H.1 Manifest ----------------------------------------
    manifest_name = manifest_sheet or _resolve_sheet(wb, ["manifest"], "--manifest-sheet")
    ws_manifest = wb[manifest_name]
    manifest_df = _ws_to_dataframe(ws_manifest, has_header=not no_header)

    return [e1_df, manifest_df]


# ---------------------------------------------------------------------------
# output helpers
# ---------------------------------------------------------------------------

_FORMATS = ("table", "csv", "tsv", "json", "markdown")


def _render(df: pd.DataFrame, fmt: str, max_rows: Optional[int]) -> str:
    if max_rows:
        df = df.head(max_rows)
    if fmt == "csv":
        return df.to_csv(index=False)
    if fmt == "tsv":
        return df.to_csv(index=False, sep="\t")
    if fmt == "json":
        return df.to_json(orient="records", indent=2)
    if fmt == "markdown":
        return df.to_markdown(index=False)
    # default: aligned table
    return df.to_string(index=False)


def _save(df: pd.DataFrame, path: pathlib.Path, fmt: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if fmt == "csv":
        df.to_csv(path, index=False)
    elif fmt == "tsv":
        df.to_csv(path, index=False, sep="\t")
    elif fmt == "json":
        df.to_json(path, orient="records", indent=2)
    elif fmt == "markdown":
        path.write_text(df.to_markdown(index=False) or "")
    else:
        path.write_text(df.to_string(index=False))
    print(f"  Saved → {path}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="extractor",
        description=(
            "Extract 'E.1. Table 5' and the Manifest sheet from an "
            ".xlsm/.xlsx workbook."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
examples:
  # print both tables to the terminal
  python extractor.py report.xlsm

  # list sheet names only
  python extractor.py report.xlsm --list-sheets

  # show only the E1 table, limited to 10 rows
  python extractor.py report.xlsm --show e1 --max-rows 10

  # export both tables as CSV files into ./output/
  python extractor.py report.xlsm --out-dir ./output --format csv

  # custom sheet names and table label
  python extractor.py report.xlsm \\
      --e1-sheet "My Norms" \\
      --manifest-sheet "My Manifest" \\
      --table-label "Table 5"

  # export single table to a specific file
  python extractor.py report.xlsm --show manifest --out-file manifest.csv --format csv
        """,
    )

    p.add_argument("file", help="Path to the .xlsm / .xlsx workbook.")

    # --- sheet / label overrides ---
    g = p.add_argument_group("sheet / table options")
    g.add_argument(
        "--e1-sheet",
        metavar="NAME",
        help="Exact name of the E1 Norm+Pool sheet (default: auto-detected).",
    )
    g.add_argument(
        "--manifest-sheet",
        metavar="NAME",
        help="Exact name of the Manifest sheet (default: auto-detected).",
    )
    g.add_argument(
        "--table-label",
        metavar="TEXT",
        default="E.1. Table 5",
        help="Substring to search for when locating the table (default: 'E.1. Table 5').",
    )
    g.add_argument(
        "--no-header",
        action="store_true",
        help="Treat the first data row as data rather than column headers.",
    )

    # --- output options ---
    o = p.add_argument_group("output options")
    o.add_argument(
        "--show",
        choices=["e1", "manifest", "both"],
        default="both",
        metavar="TABLE",
        help="Which table to display / export: e1 | manifest | both (default: both).",
    )
    o.add_argument(
        "--format", "-f",
        choices=_FORMATS,
        default="table",
        metavar="FMT",
        help=f"Output format: {' | '.join(_FORMATS)} (default: table).",
    )
    o.add_argument(
        "--max-rows",
        type=int,
        metavar="N",
        help="Limit printed/exported rows to N (useful for previewing).",
    )
    o.add_argument(
        "--out-dir",
        metavar="DIR",
        help="Save output files to this directory instead of printing to stdout.",
    )
    o.add_argument(
        "--out-file",
        metavar="FILE",
        help=(
            "Save a single table to this exact file path "
            "(only valid when --show is not 'both')."
        ),
    )

    # --- info commands ---
    i = p.add_argument_group("info commands")
    i.add_argument(
        "--list-sheets",
        action="store_true",
        help="Print all sheet names in the workbook and exit.",
    )
    i.add_argument(
        "--info",
        action="store_true",
        help="Print shape and column names for each extracted table, then exit.",
    )

    return p


def main(argv=None):
    parser = _build_parser()
    args = parser.parse_args(argv)

    # ---- info-only: list sheets ----------------------------------------
    if args.list_sheets:
        wb = openpyxl.load_workbook(args.file, read_only=True, keep_vba=False)
        print("Sheets in workbook:")
        for i, name in enumerate(wb.sheetnames, 1):
            print(f"  {i:2d}. {name}")
        return

    # ---- extract -------------------------------------------------------
    try:
        e1_df, manifest_df = extract(
            args.file,
            e1_sheet=args.e1_sheet,
            manifest_sheet=args.manifest_sheet,
            table_label=args.table_label,
            no_header=args.no_header,
        )
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        sys.exit(1)

    # ---- info summary --------------------------------------------------
    if args.info:
        for label, df in [("E.1. Table 5", e1_df), ("H.1 Manifest", manifest_df)]:
            print(f"[{label}]  shape={df.shape}")
            print(f"  columns: {list(df.columns)}\n")
        return

    # ---- validate --out-file usage ------------------------------------
    if args.out_file and args.show == "both":
        parser.error("--out-file requires --show e1 or --show manifest (not 'both')")

    tables = []
    if args.show in ("e1", "both"):
        tables.append(("E.1. Table 5", "e1_table5", e1_df))
    if args.show in ("manifest", "both"):
        tables.append(("H.1 Manifest", "manifest", manifest_df))

    # ---- save to files -------------------------------------------------
    if args.out_dir or args.out_file:
        ext_map = {"table": "txt", "csv": "csv", "tsv": "tsv",
                   "json": "json", "markdown": "md"}
        ext = ext_map[args.format]

        for label, stem, df in tables:
            if args.out_file:
                dest = pathlib.Path(args.out_file)
            else:
                dest = pathlib.Path(args.out_dir) / f"{stem}.{ext}"
            print(f"Exporting [{label}] ...")
            _save(df, dest, args.format)
        return

    # ---- print to stdout -----------------------------------------------
    for label, _, df in tables:
        print(f"=== {label} ===")
        print(_render(df, args.format, args.max_rows))
        print()


if __name__ == "__main__":
    main()
